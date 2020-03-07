
from openpyxl import load_workbook
from epidemic_models.utils.package_data import dataRootDir
import os
import numpy as np
import csv
import datetime


def restoreCovid(region='Hubei', fromItem=0):
    ''' Data from https://www.kaggle.com/sudalairajkumar/novel-corona-virus-2019-dataset#covid_19_data.csv
    '''
    rootDir = dataRootDir()
    filename = os.path.join(rootDir,
                            'covid_19_data.xlsx')

    wb = load_workbook(filename=filename)
    ws = wb[region]
    dead = np.array([c.value for c in ws['D'][fromItem:]])
    confirmed = np.array([c.value for c in ws['C'][fromItem:]])
    recovered = np.array([c.value for c in ws['E'][fromItem:]])
    date = np.array([c.value for c in ws['B'][fromItem:]])
    return dead, confirmed, recovered, date


class CovidTimeSeries():

    def __init__(self, key, state, country, lat, long, values, dates):
        self.state = state
        self.country = country
        self.latitude = lat
        self.longitude = long
        self.values = values
        self.dates = dates
        self.key = key

    @property
    def days(self):
        jan1 = datetime.datetime(2020, 1, 1, 0, 0)
        return np.array([(k - jan1).days for k in self.dates])

    @staticmethod
    def fromCSV(row, headers):
        key = row[0] + '/' + row[1] if row[0] != '' else row[1]
        dates = np.array([datetime.datetime.strptime(h, '%m/%d/%y')
                          for h in headers[4:]])
        dd = CovidTimeSeries(
            key, str(row[0]), str(row[1]),
            float(row[2]), float(row[3]),
            np.array(row[4:]).astype(float), dates)
        return dd


def _restoreGenericCSSE(filename):
    with open(filename, newline='') as csvfile:
        rd = csv.reader(csvfile, delimiter=',')
        headers = next(rd, None)
        d = {}
        for row in rd:
            ts = CovidTimeSeries.fromCSV(row, headers)
            d[ts.key] = ts
    return d


def restoreCSSEDeath():

    rootDir = dataRootDir()
    filename = os.path.join(rootDir,
                            'COVID-19',
                            'csse_covid_19_data',
                            'csse_covid_19_time_series',
                            'time_series_19-covid-Deaths.csv')
    return _restoreGenericCSSE(filename)


def restoreCSSEConfirmed():

    rootDir = dataRootDir()
    filename = os.path.join(rootDir,
                            'COVID-19',
                            'csse_covid_19_data',
                            'csse_covid_19_time_series',
                            'time_series_19-covid-Confirmed.csv')
    return _restoreGenericCSSE(filename)


def restoreCSSERecovered():

    rootDir = dataRootDir()
    filename = os.path.join(rootDir,
                            'COVID-19',
                            'csse_covid_19_data',
                            'csse_covid_19_time_series',
                            'time_series_19-covid-Recovered.csv')
    return _restoreGenericCSSE(filename)
